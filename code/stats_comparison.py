
import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont


def truncate_to_4_decimals(value):
    if not isinstance(value, (int, float)) or value is None:
        return value

    s = str(value)
    if '.' not in s:
        return value

    partie_decimale = s.split('.')[1]
    nb_decimales = len(partie_decimale)

    if nb_decimales > 4:
        return round(value, 4)
    else:
        return value


def format_val_richtext(val1, val2, color1="FFFF8C00", color2="FF3399FF"):
    if val1 is None:
        val1_str = "None"
    else:
        val1_str = f"{truncate_to_4_decimals(val1)}"

    if val2 is None:
        val2_str = "None"
    else:
        val2_str = f"{truncate_to_4_decimals(val2)}"

    tb1 = TextBlock(
        text=val1_str,
        font=InlineFont(color=color1)
    )
    tb2 = TextBlock(
        text=" / ",
        font=InlineFont(color="FF000000")
    )
    tb3 = TextBlock(
        text=val2_str,
        font=InlineFont(color=color2)
    )

    rt = CellRichText()
    rt.append(tb1)
    rt.append(tb2)
    rt.append(tb3)

    return rt


def stats_comparison_xlsx(input_json_file_1, input_json_file_2, output_file="stats_result.xlsx"):
    with open(input_json_file_1, "r", encoding="utf-8") as f1:
        json_data_1 = json.load(f1)
    with open(input_json_file_2, "r", encoding="utf-8") as f2:
        json_data_2 = json.load(f2)
    sub_flows_1 = json_data_1.get("sub-flows", [])
    sub_flows_2 = json_data_2.get("sub-flows", [])

    sub_flows_1_sorted = sorted(sub_flows_1, key=lambda f: int(f.get("label", 0)))
    sub_flows_2_sorted = sorted(sub_flows_2, key=lambda f: int(f.get("label", 0)))

    wb = Workbook()
    ws = wb.active
    ws.title = "Stats"

    row_cursor = 1
    ws.cell(row=row_cursor, column=1, value="Inter-Packet Times Stats")
    row_cursor += 1

    headers = [
        "Label",
        "min",
        "max",
        "mean",
        "mean_of_squares",
        "stddev",
        "variance",
        "coef_of_variation"
    ]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=row_cursor, column=col_idx, value=header)
    row_cursor += 1

    for flow1, flow2 in zip(sub_flows_1_sorted, sub_flows_2_sorted):
        label1 = flow1.get("label", "N/A")
        ipt1 = flow1.get("inter-packet-times", {})
        ipt2 = flow2.get("inter-packet-times", {})

        ws.cell(row=row_cursor, column=1, value=str(label1))
        ws.cell(row=row_cursor, column=2).value = format_val_richtext(ipt1.get("min"), ipt2.get("min"))
        ws.cell(row=row_cursor, column=3).value = format_val_richtext(ipt1.get("max"), ipt2.get("max"))
        ws.cell(row=row_cursor, column=4).value = format_val_richtext(ipt1.get("mean"), ipt2.get("mean"))
        ws.cell(row=row_cursor, column=5).value = format_val_richtext(ipt1.get("mean_of_squares"), ipt2.get("mean_of_squares"))
        ws.cell(row=row_cursor, column=6).value = format_val_richtext(ipt1.get("stddev"), ipt2.get("stddev"))
        ws.cell(row=row_cursor, column=7).value = format_val_richtext(ipt1.get("variance"), ipt2.get("variance"))
        ws.cell(row=row_cursor, column=8).value = format_val_richtext(ipt1.get("coef_of_variation"), ipt2.get("coef_of_variation"))

        row_cursor += 1

    row_cursor += 1

    ws.cell(row=row_cursor, column=1, value="Packet Sizes Stats")
    row_cursor += 1

    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=row_cursor, column=col_idx, value=header)
    row_cursor += 1

    for flow1, flow2 in zip(sub_flows_1_sorted, sub_flows_2_sorted):
        label1 = flow1.get("label", "N/A")
        pkt1 = flow1.get("packet-sizes", {})
        pkt2 = flow2.get("packet-sizes", {})

        ws.cell(row=row_cursor, column=1, value=str(label1))
        ws.cell(row=row_cursor, column=2).value = format_val_richtext(pkt1.get("min"), pkt2.get("min"))
        ws.cell(row=row_cursor, column=3).value = format_val_richtext(pkt1.get("max"), pkt2.get("max"))
        ws.cell(row=row_cursor, column=4).value = format_val_richtext(pkt1.get("mean"), pkt2.get("mean"))
        ws.cell(row=row_cursor, column=5).value = format_val_richtext(pkt1.get("mean_of_squares"), pkt2.get("mean_of_squares"))
        ws.cell(row=row_cursor, column=6).value = format_val_richtext(pkt1.get("stddev"), pkt2.get("stddev"))
        ws.cell(row=row_cursor, column=7).value = format_val_richtext(pkt1.get("variance"), pkt2.get("variance"))
        ws.cell(row=row_cursor, column=8).value = format_val_richtext(pkt1.get("coef_of_variation"), pkt2.get("coef_of_variation"))

        row_cursor += 1

    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 25
        for row in range(1, row_cursor + 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(output_file)
